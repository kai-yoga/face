import os
# import json
import openpyxl as opl
import face_recognition
from datetime import datetime
from openpyxl.drawing.image import Image
# import getconfig

def getToConvert(libs, others):
    '''
    判断others与
    :param libs:
    :param others:
    :return:返回others对应的libs中的文件名
    '''
    libs_file_key = libs["file_key"]
    to_convert = (0, 0)
    file_key = others["file_key"]
    if libs_file_key.lower() != file_key.lower():
        if libs_file_key.lower() == "xh" and file_key == "zjhm":
            '''标准文件为学号，当前待比较文件为证件号码'''
            to_convert = (1, 0)
        elif libs_file_key.lower() == "zjhm" and file_key == "xh":
            '''标准文件为证件号码，当前待比较文件为学号'''
            to_convert = (0, 1)
    return to_convert


def getConvertFile(convert_file_path):
    '''读取转换文件，转换文件第一列是学号，第二列是证件号码'''
    if os.path.exists(convert_file_path):
        wb = opl.load_workbook(convert_file_path)
        L = []
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            t = (row[0], row[1])
            L.append(t)
        return L


def getOthers(path):
    '''
    返回待比较文件的生成器
    :param paths: 待比较文件的路径列表
    :return: 文件名
    '''
    if os.path.exists(path):
        for file in os.listdir(path):
            yield file


def getLibFileName(L, file, to_convert):
    '''
    返回标准库中的文件名
    :param L:转换文件的内容
    :param file:待比较的文件名
    :param to_convert:转换规则
    :return:返回标准库中的文件名
    '''
    if L is not None and len(L) >= 1:
        for ele in L:
            if ele[to_convert[0]] == file:
                return ele[to_convert[1]]
    else:
        return file


def getSimilar(lib_file, other_file):
    '''
    :param lib_file: 标准库中的文件
    :param other_file: 待比较的文件
    :return:返回两个文件的相似度
    '''
    lib_face = face_recognition.load_image_file(lib_file)
    lib_encoding = face_recognition.face_encodings(lib_face)[0]
    other_face = face_recognition.load_image_file(other_file)
    other_encoding = face_recognition.face_encodings(other_face)[0]
    similar = 1 - face_recognition.face_distance([lib_encoding],
        other_encoding)[0]
    return round(similar, 4)


def expxlsx(res_file_name, res, type):
    '''输出比对结果
    :param res_file_name:保存的文件名
    :param res:要保存的内容
    :param type:要保存的文件类型
    '''
    if res != [('比较', '基础', '相似度', '比较照片', '基础照片')]:
        if type.lower() == 'xlsx':
            wb = opl.Workbook()
            ws = wb.active
            ws.append(('比较', '基础', '相似度', '比较照片', '基础照片'))
            for i in range(1, len(res)):
                ws.append(res[i])
                ws.row_dimensions[i + 1].height = 50
                ws.column_dimensions['D'].width = 50
                ws.column_dimensions['E'].width = 50
                image = Image(res[i][0])
                image.height = 50
                image.width = 50
                ws.add_image(image, 'D' + str(i + 1))
                try:
                    image = Image(res[i][1])
                    image.height = 50
                    image.width = 50
                    ws.add_image(image, 'E' + str(i + 1))
                except:
                    ws.cell(row=i + 1, column=5).value = res[i][1]

            wb.save(res_file_name)
        elif type.lower() == 'txt':
            with open(res_file_name, 'w', encoding='utf-8') as f:
                if res != [('比较', '基础', '相似度', '比较照片', '基础照片')]:
                    for row in res:
                        f.write(','.join(list(row)) + '\n')
                else:
                    f.write('没有检测到相似度低于指定的similar的照片！')
            f.close()
        else:
            pass


def main(config):
    '''
    主函数
    :return:
    '''
    config = config["bdzp"]
    if config["open"] not in ["0"]:
        compare = config["compare"]
        libs = config["libs"]
        if libs["path"] and libs["file_key"]:
            res = [('比较', '基础', '相似度', '比较照片', '基础照片')]
            for others in config["others"]:
                if others["file_key"] and others["path"] and others["similar"]:
                    to_convert = getToConvert(libs, others)
                    convert_file = None
                    if to_convert != (0, 0):
                        convert_file = getConvertFile(others["convert_file_path"])
                    for file in getOthers(others["path"]):
                        print('current {}'.format(os.path.join(others["path"],file)))
                        lib_file = getLibFileName(convert_file, file, to_convert)
                        other_file = os.path.join(others["path"], file)
                        similar = 0
                        if os.path.exists(os.path.join(libs["path"], lib_file)):  ###lib_file中不存在对应的文件
                            lib_file = os.path.join(libs["path"], lib_file)
                            if lib_file.split('.')[-1] in compare["suffix"] and other_file.split('.')[-1] in compare["suffix"]:
                                similar = getSimilar(lib_file, other_file)
                        else:
                            lib_file = '不存在'
                            # similar = 0
                        if similar < float(others["similar"]):
                            res.append((other_file, lib_file, str(similar)))
                res_file_name = os.path.join(compare["save_path"],
                                             datetime.now().strftime('%Y-%m-%d %H%M%S') + '.' + compare["outputFileType"].lower())
                expxlsx(res_file_name, res, compare["outputFileType"].lower())


# if __name__ == "__main__":
#     main()
